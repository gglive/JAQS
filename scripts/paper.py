import sys, os
import datetime, time
from collections import defaultdict 

import pymysql
import sqlalchemy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from WindPy import w

w.start ()
today = datetime.datetime.today().strftime("%Y-%m-%d")
prev_tday = w.tdaysoffset (-1, today, "").Times[0]
print (today)
print (prev_tday)


# Connect to the database
mydb = pymysql.connect(
    host='1.85.40.235',
    port=28888,
    user='kods',
    password='Abc1234#',
    db='strategy_kaiyuan',
    cursorclass=pymysql.cursors.DictCursor
)

## Trade Book
workbook = load_workbook ( filename=u"C://Users/admin/Documents/alpha/_Trading.xlsx", data_only=True )
##
## 策略组合
portfolios = {}
plist = {}
##
worksheet = workbook.get_sheet_by_name ('params')
values = worksheet.values
cols = next ( values)
rows = list ( values)
for row in rows:
    p = {}
    for n, col in enumerate(cols):
        p[col] = row[n]
    
    tag = p['tag'] # row[0]
    print ('Portfolio', tag)
    plist[tag] = p

if len (sys.argv) > 1:
    for p in sys.argv[1:]: portfolios [p] = plist[p]
else:
    portfolios = plist


## 前日持仓
positions = defaultdict(dict)
with mydb.cursor() as cursor:
    # Read a single record
    sql = "SELECT * FROM `_legacy_position_2` WHERE `trade_day`=%s AND `paper_flag`=%s AND `hedge_flag`=%s AND `held_qty`>0"
    cursor.execute(sql, (prev_tday, 'P', ''))
    result = cursor.fetchall ()
    for row in result:
        tag = row['tag']
        key = row['security_id']
        positions[tag][key] = row

## 导入订单
orders = []
###
for key, p in portfolios.items():
    print ( key)
    worksheet = workbook.get_sheet_by_name ( key)
    print ( "Row:", worksheet.min_row, worksheet.max_row)
    print ( "Col:", worksheet.min_column, worksheet.max_column)
    values = worksheet.values
    cols = next ( values)
    rows = list ( values) 
    if worksheet.max_row < 2:
        continue
    for row in rows:
        o = {}
        for n, col in enumerate(cols):
            o[col] = row[n]
        ##
        # o['security_id'] = f"{o['security_id']:06d}"
        o['security_id_impl'] = o['security_id']# f"{o['security_id']}.{o['exchange_id']}"
        orders.append( o)

i = 0
for order in orders:
    ###
    # 准备订单
    #
    tag = order['order_tag']
    key = order['security_id_impl'] # f"{order['security_id']}.{order['exchange_id']}"
    if order ['order_bsFlag'] == 'S':
        if order['order_qty'] == 0:
            position = positions[tag].get ( key, { 'security_id': key, 'held_qty': 0} )
            order['order_qty'] = position['held_qty']
        else:
            # order['order_qty'] = 
            pass
    else:
        portfolio = portfolios [tag]
        if order['order_qty'] == 0:
            order['order_qty'] = int (round( portfolio['s'] / order['offer_refPx']/100,0 ) * 100)
        else:
            # order['order_qty'] =
            pass
    
    ###
    i += 1
    print (f"Proposal Order {i}: {order['order_bsFlag']} - Security:{order['security_id_impl']}, Qty:{order['order_qty']}, {order['order_tag']}")

i = 0

### 
# 执行订单
#
with mydb.cursor() as cursor:
    sql = """
            INSERT INTO 
                `_legacy_knock_2` 
                (`trade_day`, `acct_id`,`security_id`, `order_id`, `order_side`, `order_avgpx`, `order_qty`,`fee`, `hedge_flag`, `paper_flag`, `tag`)
            VALUES 
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
    for order in orders:
        if order['offer_avgPx'] <= 0:
            print (order["security_id"], "Can't trade in/out - Stock suspended ")
            order['order_qty'] = 0

        if order['order_bsFlag'] == 'B' and order['offer_avgPx'] >= order['px_lmt_up'] :
            print (order["security_id"], "Can't buy in - Reach up limit")
            order['order_qty'] = 0
        if order['order_bsFlag'] == 'B' and order.get('name', '').startswith('ST'):
            print (order['security_id'], "Can't buy in - Stock with `ST`")

        if order['order_bsFlag'] == 'S' and order['offer_avgPx'] <= order['px_lmt_down']:
            print (order["security_id"], "Can't sell out - Reach down limit")
            order['order_qty'] = 0

        i += 1
        order['order_id'] = f"#{i:04d}"

        # if order['order_qty'] == 0:
        #     continue
        order['order_amount'] = order['offer_avgPx'] * order['order_qty']
        order['fee'] =round(order['order_amount']*0.0000487, 2) + round(order['order_amount']*0.00002, 2) + round(order['order_amount']*0.00002, 2)
        if order['order_bsFlag'] == 'S':
            order['fee'] += round(order['order_amount'] *0.001, 2)

        # single record
        cursor.execute( sql, 
        (   order['trade_day'], 
            order.get('acct_id', "999"),
            order['security_id_impl'], # f"{order['security_id']}.{order['exchange_id']}",
            order['order_id'],
            order['order_bsFlag'],
            order['offer_avgPx'],
            order['order_qty'], 
            order['fee'],
            '',
            'P',
            order['order_tag']
        ))
        print (f"Place Order: Id:{order['order_id']} - Security:{order['security_id_impl']}, Qty:{order['order_qty']}, {order['order_tag']}")
    
    mydb.commit ()
