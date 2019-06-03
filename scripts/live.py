# encoding: UTF-8

import sys, os
from datetime import datetime, timedelta
from collections import defaultdict 

from jaqs.trade import model
from jaqs.trade.tradegateway import RealTimeTradeApi
from jaqs.data.dataservice import RemoteDataService

import pymysql
import sqlalchemy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from WindPy import w


w.start ()
today = datetime.today().strftime("%Y-%m-%d")
prev_tday = w.tdaysoffset (-1, today, "").Times[0]
print (today)
print (prev_tday)

# Connect to the database
mydb = pymysql.connect(
    host='1.85.40.235',
    port=28888,
    user='kods',
    password='Abc1234#',
    db='strategy_kaiyuan',
    cursorclass=pymysql.cursors.DictCursor
)


## Trade Book
workbook = load_workbook ( filename=u"C://Users/admin/Documents/alpha/_Trading.xlsx", data_only=True )
##
## 策略组合
portfolios = {}
plist = {}
##
worksheet = workbook.get_sheet_by_name ('params')
values = worksheet.values
cols = next ( values)
rows = list ( values)
for row in rows:
    p = {}
    for n, col in enumerate(cols):
        p[col] = row[n]
    
    tag = p['tag'] # row[0]
    print ('Portfolio', tag)
    plist[tag] = p

if len (sys.argv) > 1:
    for p in sys.argv[1:]: portfolios [p] = plist[p]
else:
    portfolios = plist

## 前日持仓
positions = defaultdict(dict)
with mydb.cursor() as cursor:
    # Read a single record
    sql = "SELECT * FROM `_legacy_position_2` WHERE `trade_day`=%s AND `paper_flag`=%s AND `hedge_flag`=%s AND `held_qty`>0"
    cursor.execute(sql, (prev_tday, 'T', ''))
    result = cursor.fetchall ()
    for row in result:
        tag = row['tag']
        key = row['security_id']
        positions[tag][key] = row

## 导入订单
orders = []
###
for key, p in portfolios.items():
    #if p['activity'] == "P":
    #    continue

    print ( key)
    worksheet = workbook.get_sheet_by_name ( key)
    print ( "Row:", worksheet.min_row, worksheet.max_row)
    print ( "Col:", worksheet.min_column, worksheet.max_column)
    values = worksheet.values
    cols = next ( values)
    rows = list ( values) 
    if worksheet.max_row < 2:
        continue
    for row in rows:
        o = {}
        for n, col in enumerate(cols):
            o[col] = row[n]
        ##
        #o['security_id'] = f"{o['security_id']:06d}"
        o['security_id_impl'] = o['security_id']#f"{o['security_id']}.{o['exchange_id']}"
        orders.append( o)

i = 0
for order in orders:
    ###
    # 准备订单
    #
    tag = order['order_tag']
    key = order['security_id_impl']
    if order ['order_bsFlag'] == 'S':
        if order['order_qty'] == 0:
            position = positions[tag].get ( key, { 'security_id': key, 'held_qty': 0} )
            order['order_qty'] = position['held_qty']
        else:
            # order['order_qty'] = 
            pass
        order ['order_side_impl'] = 'Sell'
    else:
        if order['order_qty'] == 0:
            portfolio = portfolios [tag]
            order['order_qty'] = int (round( portfolio['s'] / order['offer_refPx']/100,0 ) * 100)
        else:
            pass
        order ['order_side_impl'] = 'Buy'
    ###
    i += 1
    print (f"Proposal Order {i}: {order['order_bsFlag']} - Security:{order['security_id']}, Qty:{order['order_qty']}, {order['order_tag']}")



yesOrNo = input("确认下单（Y/N）：")
if yesOrNo.upper() != 'Y':
    sys.exit(0)

### 
# 执行订单
#
trade_api = RealTimeTradeApi({
    "remote.trade.address": "tcp://1.85.40.235:58086",
    "remote.trade.username": 'diryox',
    #"remote.trade.address": "tcp://127.0.0.1:58086",
    #"remote.trade.username": "rqalpha",
    "remote.trade.password": "10086"
})
trade_api.init_from_config({
    "remote.trade.address": "tcp://1.85.40.235:58086",
    "remote.trade.username": "diryox",
    #"remote.trade.address": "tcp://127.0.0.1:58086",
    #"remote.trade.username": "rqalpha",
    "remote.trade.password": "10086"
})
context = model.Context(data_api=None, trade_api=trade_api, instance=None, strategy=None, pm=None)

#
for order in orders:
    if order['order_qty'] == 0:
        continue

    if order['order_place_method'] == "TWAP":
        offer_start_time = order['offer_start_time'].strftime("%H:%M:%S")
        offer_stop_time = order['offer_stop_time'].strftime("%H:%M:%S")
    elif order['order_place_method'] == "DMA":
        now  = datetime.now()
        _date = now.date()
        _time = now.time()
        _time = max ( _time, order['offer_start_time'] )
        offer_start_time = _time.strftime("%H:%M:%S")
        offer_stop_time = (datetime.combine(_date, _time) + timedelta(seconds=60)).strftime("%H:%M:%S")
    else:
        print ("Unknown place method, ", order['security_id_impl'], order['order_place_method'] )
        continue

    order_id, msg = trade_api.place_order( 
        order['security_id_impl'], order['order_side_impl'], 0, order['order_qty'],
        algo='TWAP_KY_01', 
        algo_param={ 
            'algo.style': 2,
            'algo.order_position': 'OP1',
            'algo.order_tick': 99 ,
            'algo.append_position': 'OP1',
            'algo.append_tick': 99,
            'algo.cancel_cycle': 60,
            'start_time': offer_start_time,
            'stop_time':  offer_stop_time,
        },
        userdata=order['order_tag']
    )
    print (f"Order: Id:{order_id} - Security:{order['security_id_impl']}, Qty:{order['order_qty']}, {order['order_tag']}")