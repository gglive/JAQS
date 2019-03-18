# encoding: UTF-8

import sys, os
import datetime, time
from collections import defaultdict 

from jaqs.trade import model
from jaqs.trade.tradegateway import RealTimeTradeApi
from jaqs.data.dataservice import RemoteDataService

import pymysql
import sqlalchemy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from WindPy import w


mode = input("Live or Paper（L/P）：")
mode = mode.upper()

w.start ()

today = datetime.datetime.today().strftime("%Y-%m-%d")
print (today)

prev_tday = w.tdaysoffset (-1, today, "").Times[0]
print (prev_tday)

# Connect to the database
connection = pymysql.connect(
    host='1.85.40.235',
    port=28888,
    user='kods',
    password='Abc1234#',
    db='strategy_kaiyuan',
    cursorclass=pymysql.cursors.DictCursor
)

portfolios = {}

holdings = {}
positions = defaultdict (dict)


with connection.cursor() as cursor:
    sql = "SELECT * FROM `_legacy_profit_n_loss_2` WHERE `trade_day`=%s"
    cursor.execute(sql, (prev_tday,))
    result = cursor.fetchall ()
    for record in result:
        portfolios[ f"{record['tag']}-{record['sp_0']}-{record['sp_1']}"] = record

with connection.cursor() as cursor:
    # Read a single record
    sql = "SELECT * FROM `_legacy_position_2` WHERE `trade_day`=%s AND `tag` !=\'\'"
    cursor.execute(sql, (prev_tday,))
    result = cursor.fetchall ()
    for record in result:
        holdings[record['security_id']] = { "security_id": record['security_id'] } 
        positions[f"{record['tag']}-{record['hedge_flag']}-{record['paper_flag']}"][record['security_id']] = record
                  

orders = []

#import sys
#if argv[1].upper() == '-R':
    
workbook = load_workbook ( filename=u"C://Users/admin/Documents/alpha/_Trading.xlsx", data_only=True )

###
worksheet = workbook.get_sheet_by_name ('weights')
values = worksheet.values
cols = next ( values)
rows = list ( values) 

for row in rows:
    if mode != 'L':
        p = portfolios [row[0] + '--P']
    else:
        p = portfolios [row[0] + '--T']

    p['c'] = row[1]
    p['w'] = row[2]
    p['w0'] = row[3]
    p['s'] = row[4]
    p['op'] = row[5]


worksheet = workbook.get_sheet_by_name ("orders")
print ( worksheet.min_row, worksheet.max_row)
print ( worksheet.min_column, worksheet.max_column)
values = worksheet.values
cols = next ( values)
rows = list ( values) 

for row in rows:
    cmd = {}
    for n, col in enumerate(cols):
        cmd[col] = row[n]
    
    cmd['paper_flag'] = 'P' if mode != 'L' else 'T'
    holdings[cmd['security_id']] = { "security_id": cmd['security_id'] } 
    
    orders.append( cmd)



wData = w.wss(",".join(holdings.keys()), "sec_name,pre_close,close",f"tradeDate={today};priceAdj=U;cycle=D")
for n, code in enumerate(wData.Codes):
    hold = holdings[code]
    hold['security_name'] = wData.Data[0][n]
    hold['prev_close'] = wData.Data[1][n]
    hold['close'] = wData.Data[2][n]


for order in orders:
    ###
    # 准备订单
    #
    tag = order['tag'] + '--' + order['paper_flag']
    if order ['order_side'] == 'S':
        position = positions[tag].pop (order['security_id'], { 'held_qty': 0} )
        order['order_qty'] = position['held_qty']
        order['order_side_impl'] = 'Sell'
    else:
        portfolio = portfolios[ tag ]
        hold = holdings[ order['security_id'] ]
        order['order_qty'] = int(round( portfolio['op']/ hold['prev_close']/100,0 ) * 100) # position['price_settlement']
        order['order_side_impl'] = 'Buy'
    
    ###
    print (f"Cmd: {order['order_side_impl']} - Security:{order['security_id']}, Qty:{order['order_qty']}, {order['tag']}")


# 仓位调整
for portfolio in portfolios.values():
    if not 'c' in portfolio:
        continue

    v0 = portfolio['units'] * portfolio['pct_val']
    v1 = portfolio['c'] * portfolio['w']
    
    delta = v0 - v1
    print ("Delta", delta)
    diff = abs ( delta )
    
    diff = diff / portfolio['s'] 
    print ("Diff", diff)
    pd = positions[ portfolio['tag'] ]
    for position in pd.values():

        if delta < 0:
            hold = holdings[ position['security_id'] ]
            cmd = {
                'security_id': position['security_id'],
                'order_side': 'B',
                'order_side_impl': 'Buy',
                'order_qty': int( diff / hold['prev_close'] / 100)* 100
            }
        else:
            cmd = {
                'security_id': position['security_id'],
                'order_side': 'S',
                'order_side_impl': 'Sell',
                'order_qty': int( diff / hold['prev_close'] / 100 )* 100
            }
        cmd['tag'] =  portfolio['tag'] 
        ###
        print (f"Cmd: {cmd['order_side_impl']} - Security:{cmd['security_id']}, Qty:{cmd['order_qty']}, {cmd['tag']}")
        ###
        orders.append (cmd)

# import signal

yesOrNo = input("确认下单（Y/N）：")
if yesOrNo.upper() != 'Y':
    sys.exit(0)

if mode == 'L':

    trade_api = RealTimeTradeApi({
        "remote.trade.address": "tcp://1.85.40.235:58086",
        "remote.trade.username": 'diryox',
        #"remote.trade.address": "tcp://127.0.0.1:58086",
        #"remote.trade.username": "rqalpha",
        "remote.trade.password": "10086"
    })
    trade_api.init_from_config({
        "remote.trade.address": "tcp://1.85.40.235:58086",
        "remote.trade.username": "diryox",
        # "remote.trade.address": "tcp://127.0.0.1:58086",
        # "remote.trade.username": "rqalpha",
        "remote.trade.password": "10086"
    })
    context = model.Context(data_api=None, trade_api=trade_api, instance=None, strategy=None, pm=None)
    
    #
    for order in orders:

        offer_stop_time = '14:57:00'
        
        if order['order_side'] == 'B':
            offer_start_time =  '09:30:00'
        else: #order['order_side'] == 'S'
            offer_start_time = '14:30:00'
        
        if order['order_qty'] == 0:
            continue
            
        order_id, msg = trade_api.place_order( 
            order['security_id'], order['order_side_impl'], 0, order['order_qty'],
            algo='TWAP_KY_01', 
            algo_param={ 
                'algo.style': 2,
                'algo.order_position': 'OP1',
                'algo.order_tick': 99 ,
                'algo.append_position': 'OP1',
                'algo.append_tick': 99,
                'algo.cancel_cycle': 60,
                'start_time': offer_start_time,
                'stop_time':  offer_stop_time,
            },
            userdata=order['tag']
        )
        print (f"Order: Id:{order_id} - Security:{order['security_id']}, Qty:{order['order_qty']}, {order['tag']}")

else:
    
    with connection.cursor() as cursor:
        
        sql = """
                INSERT INTO 
                    `_legacy_knock_2` 
                    (`trade_day`, `acct_id`,`security_id`, `order_id`, `order_side`, `order_avgpx`, `order_qty`,`fee`, `hedge_flag`, `paper_flag`, `tag`)
                VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """
        for order in orders:
            if order['order_side'] == 'B' and order['涨跌停'] == 'U':
                order['order_qty'] == 0
            if order['order_side'] == 'S' and order['涨跌停'] == 'D':
                order['order_qty'] == 0
            if order['trading_status'] == '未知':
                order['order_qty'] = 0

            # if order['order_qty'] == 0:
            #     continue
            order['order_avgpx'] = order['latest_px'] # holdings[ order['security_id']] ['close']
            order['order_amt_impl'] = order['order_avgpx'] * order['order_qty']
            
            order['fee'] =round(order['order_amt_impl']*0.0000487, 2) + round(order['order_amt_impl']*0.00002, 2) + round(order['order_amt_impl']*0.00002, 2)
            if order['order_side'] == 'S':
                order['fee'] += round(order['order_amt_impl'] *0.001, 2)

            # single record
            cursor.execute(sql, 
                (   today, 
                    order['acct_id'],
                    order['security_id'],
                    order['order_id'],
                    order['order_side'],
                    order['order_avgpx'],
                    order['order_qty'], 
                    order['fee'],
                    '',
                    'P',
                    order['tag'])
                )
            print (f"Order: Id:{order['order_id']} - Security:{order['security_id']}, Qty:{order['order_qty']}, {order['tag']}")
        
        connection.commit ()
